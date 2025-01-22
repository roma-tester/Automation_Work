import openpyxl


def dataGenerator():
    # li = [['user1', 'pass1'],['user2', 'pass2'],['user3', 'pass3']]
    # return li
    wk = openpyxl.load_workbook("D:/PROJECT_STRUCTURE/TData.xlsx")
    sh = wk['Sheet1']
    r = sh.max_row
    li =[]
    li1 = []
    for i in range(1, r+1):
        li1 = []
        un = sh.cell(i, 1)
        up = sh.cell(i, 2)
        li1.insert(0, un.value)
        li1.insert(1, up.value)
        li.insert(i-1, li1)

    print(li)
    return li
